import numpy as np
import openpyxl

# Constants

training_pos = 5000 # of positive training examples (upper bound)
training_neg = 4000 # of negative training examples (upper bound)
training_sheet = "TRAINING.xlsx" # training file name
training_pname = "positive" # name of training sheet w/ positive sentiments
training_nname = "negative" # ^^ but negative
testing_sheet = "TESTING.xlsx" # name of testing excel file
testing_name = "Sheet1" # testing sheet name
testing_lb = 2 # testing range lower value
testing_ub = 232360 # testing range upper value
result_col = 'I' # column to store results in testing file
data_col = 'G' # column where tweet data is stored
state_col = 'D' # column where state is stored

class Counter(dict):
    def __missing__(self, key): return 0

class DefaultDict(dict):
    def __init__(self, default_factory): self._default_factory = default_factory
    def __missing__(self, key): return self._default_factory()

def realword(w):
    if w.startswith("http"): return False #if w.startswith("@"): return False
    if w.startswith("RT"): return False
    return True

def get_counts(file_list):
    counts = Counter()
    for f in file_list:
        words = f.split()
        for w in set(words):
            if realword(w):
                counts[w] += 1
    return counts

def get_log_probabilities(file_list):
    counts = get_counts(file_list)
    N_files = len(file_list)
    N_categories = 2
    log_prob = DefaultDict(lambda : -np.log(N_files + N_categories))
    for word in counts:
        log_prob[word] = np.log(counts[word] + 1) - np.log(N_files + N_categories)
        assert log_prob[word] < 0
    return log_prob

def learn_distributions(file_lists_by_category):
    log_probs_by_category = []
    prior_by_category = []

    for file_list in file_lists_by_category:
        log_prob = get_log_probabilities(file_list)
        log_probs_by_category.append(log_prob)
        prior_by_category.append(len(file_list))

    log_prior_by_category = [np.log(p/sum(prior_by_category)) for p in prior_by_category]

    return (log_probs_by_category, log_prior_by_category)

def classify_message(message_filename,
                     log_probabilities_by_category,
                     log_prior_by_category,
                     names):
    try: message_words = set(message_filename.split())
    except: return 'none'
    
    N_categories  = len(log_probabilities_by_category)
    all_words = []
    for i in range(N_categories):
        all_words += log_probabilities_by_category[i].keys()
    all_words = list(set(all_words))

    log_likelihoods = []
    for i in range(N_categories):
        total = 0
        all_word_log_probs = log_probabilities_by_category[i]
        for w in all_words:
            log_prob = all_word_log_probs[w]
            test = (w in message_words)
            total += test*log_prob + (1-test)*np.log(1-np.exp(log_prob))
        log_likelihoods.append(total)
    posterior = np.array(log_likelihoods) + np.array(log_prior_by_category)
    winner = np.argmax(posterior)
    return names[winner]

if __name__ == '__main__':

    print("Training...")
    wb = openpyxl.load_workbook(training_sheet)
    sheet = wb[training_pname]
    file_lists = [[],[]]
    
    for i in range(1, training_pos):
        if sheet['A'+str(i)].value not in ["", None]:
            file_lists[0].append(sheet['A'+str(i)].value)

    sheet = wb[training_nname]
    for i in range(1, training_neg):
        if sheet['A'+str(i)].value not in ["", None]:
            file_lists[1].append(sheet['A'+str(i)].value)
    
    (log_probabilities_by_category, log_priors_by_category) = \
            learn_distributions(file_lists)

    print("Testing...")
    wb = openpyxl.load_workbook(testing_sheet)
    sheet = wb[testing_name]

    for i in range(testing_lb, testing_ub):
        if i % 1000 == 0: print(i)
        filename = sheet[data_col+str(i)].value
        if type(filename) == None:
            filename = ""
        label = classify_message(filename, log_probabilities_by_category,
                                log_priors_by_category, ['positive', 'negative'])
        if label == 'positive': sheet[result_col+str(i)].value = 1
        elif label == 'negative': sheet[result_col+str(i)].value = -1

    print("Analyzing Data...")
    states = {}
    for i in range(testing_lb, testing_ub):
        try:
            state = sheet[state_col+str(i)].value
            sentiment = (int(sheet[result_col+str(i)].value)+1)//2
            if state not in states: states[state] = [0, 0]
            states[state][sentiment] += 1
        except:
            continue
    
    wb.create_sheet('Stats')
    sheet = wb['Stats']
    sheet['B1'].value = "# Negative"
    sheet['C1'].value = "# Positive"
    sheet['D1'].value = "% Positive"
    row = 1
    for i in states:
        row += 1
        sheet['A'+str(row)].value = i
        sheet['B'+str(row)].value = states[i][0]
        sheet['C'+str(row)].value = states[i][1]
        sheet['D'+str(row)].value = states[i][1]/(states[i][0]+states[i][1])
    print("Saving...")
    wb.save(testing_sheet)
    print("Done!")
